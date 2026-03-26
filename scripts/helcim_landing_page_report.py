#!/usr/bin/env python3
"""
Helcim Inc. — Landing Page Conversion Rate Report
Pulls last 30 days of data from Google Ads and reports:
  1. Unique landing page URLs with clicks (normalized — {ignore} stripped)
  2. CVR: click → Conversion 1 (E-mail Captured)  [from all_conversions]
  3. CVR: click → primary conversion event         [from metrics.conversions]

Output: console + output/helcim_landing_page_report.md
"""

import os
from collections import defaultdict
from datetime import datetime
from urllib.parse import urlparse, urlunparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.ads.googleads.client import GoogleAdsClient

CUSTOMER_ID = "9954926882"
EMAIL_CAPTURED_ACTION = "Conversion 1 (E-mail Captured)"
DATE_RANGE = "LAST_30_DAYS"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
OUTPUT_MD = os.path.join(OUTPUT_DIR, "helcim_landing_page_report.md")
OUTPUT_XLS = os.path.join(OUTPUT_DIR, "helcim_landing_page_report.xlsx")


def normalize_url(url):
    """Strip {ignore}, query params, and normalize trailing slashes."""
    # Remove {ignore} from the path
    url = url.replace("{ignore}", "")
    parsed = urlparse(url)
    # Drop query string and fragment (all tracking params)
    path = parsed.path.rstrip("/") or "/"
    clean = urlunparse((parsed.scheme, parsed.netloc, path, "", "", ""))
    return clean


def cvr(conversions, clicks):
    if clicks == 0:
        return "—"
    return f"{conversions / clicks:.2%}"


def fetch_clicks(ga):
    """Query 1: clicks by landing page + campaign (no conversion segmentation)."""
    query = f"""
        SELECT
          landing_page_view.unexpanded_final_url,
          campaign.name,
          campaign.status,
          metrics.clicks
        FROM landing_page_view
        WHERE campaign.status = 'ENABLED'
          AND segments.date DURING {DATE_RANGE}
          AND metrics.clicks > 0
        ORDER BY metrics.clicks DESC
    """
    data = defaultdict(int)
    for row in ga.search(customer_id=CUSTOMER_ID, query=query):
        url = normalize_url(row.landing_page_view.unexpanded_final_url)
        campaign = row.campaign.name
        data[(url, campaign)] += row.metrics.clicks
    return data


def fetch_conversions(ga):
    """Query 2: conversions by landing page + campaign + conversion action."""
    query = f"""
        SELECT
          landing_page_view.unexpanded_final_url,
          campaign.name,
          campaign.status,
          segments.conversion_action_name,
          metrics.all_conversions,
          metrics.conversions
        FROM landing_page_view
        WHERE campaign.status = 'ENABLED'
          AND segments.date DURING {DATE_RANGE}
          AND metrics.all_conversions > 0
    """
    data = defaultdict(lambda: {"email": 0.0, "primary": 0.0})
    for row in ga.search(customer_id=CUSTOMER_ID, query=query):
        url = normalize_url(row.landing_page_view.unexpanded_final_url)
        campaign = row.campaign.name
        action = row.segments.conversion_action_name
        key = (url, campaign)
        if action == EMAIL_CAPTURED_ACTION:
            data[key]["email"] += row.metrics.all_conversions
        data[key]["primary"] += row.metrics.conversions
    return data


def build_rows(clicks_data, conv_data):
    """Merge clicks and conversions into per-(url, campaign) rows."""
    all_keys = set(clicks_data.keys()) | set(conv_data.keys())
    rows = []
    for key in all_keys:
        url, campaign = key
        clicks = clicks_data.get(key, 0)
        email = conv_data[key]["email"] if key in conv_data else 0.0
        primary = conv_data[key]["primary"] if key in conv_data else 0.0
        rows.append({
            "url": url,
            "campaign": campaign,
            "clicks": clicks,
            "email": email,
            "primary": primary,
        })
    rows.sort(key=lambda r: r["clicks"], reverse=True)
    return rows


def summarize_by_url(rows):
    """Roll up rows to landing page level."""
    by_url = defaultdict(lambda: {"clicks": 0, "email": 0.0, "primary": 0.0})
    for r in rows:
        by_url[r["url"]]["clicks"] += r["clicks"]
        by_url[r["url"]]["email"] += r["email"]
        by_url[r["url"]]["primary"] += r["primary"]
    summary = [{"url": url, **vals} for url, vals in by_url.items()]
    summary.sort(key=lambda r: r["clicks"], reverse=True)
    return summary


def summarize_by_url_campaign(rows):
    """Roll up rows to (landing page, campaign) level."""
    by_key = defaultdict(lambda: {"clicks": 0, "email": 0.0, "primary": 0.0})
    for r in rows:
        key = (r["url"], r["campaign"])
        by_key[key]["clicks"] += r["clicks"]
        by_key[key]["email"] += r["email"]
        by_key[key]["primary"] += r["primary"]
    result = [{"url": url, "campaign": camp, **vals} for (url, camp), vals in by_key.items()]
    # Sort by url (by total clicks desc across all campaigns), then campaign
    url_clicks = defaultdict(int)
    for r in result:
        url_clicks[r["url"]] += r["clicks"]
    result.sort(key=lambda r: (-url_clicks[r["url"]], r["url"], -r["clicks"]))
    return result


def md_table_summary(summary):
    lines = [
        "| Landing Page | Clicks | Email Captured | Email CVR | Primary Conv | Primary CVR |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for r in summary:
        lines.append(
            f"| {r['url']} | {r['clicks']:,} | {r['email']:.1f} | {cvr(r['email'], r['clicks'])} "
            f"| {r['primary']:.1f} | {cvr(r['primary'], r['clicks'])} |"
        )
    return "\n".join(lines)


def md_table_by_url_campaign(rows):
    """Table grouped by landing page with campaigns as sub-rows."""
    lines = [
        "| Landing Page | Campaign | Clicks | Email Captured | Email CVR | Primary Conv | Primary CVR |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    current_url = None
    for r in rows:
        url_cell = r["url"] if r["url"] != current_url else "↳"
        current_url = r["url"]
        lines.append(
            f"| {url_cell} | {r['campaign']} | {r['clicks']:,} | {r['email']:.1f} | {cvr(r['email'], r['clicks'])} "
            f"| {r['primary']:.1f} | {cvr(r['primary'], r['clicks'])} |"
        )
    return "\n".join(lines)


def write_xlsx(summary, by_url_campaign, run_date):
    HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
    URL_FILL = PatternFill("solid", fgColor="E8EAF6")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

    def pct_val(conv, clicks):
        return (conv / clicks) if clicks else None

    def add_header(ws, cols):
        for i, (label, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = width

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary by Landing Page ---
    ws1 = wb.active
    ws1.title = "Summary by Landing Page"
    cols1 = [
        ("Landing Page", 55),
        ("Clicks", 10),
        ("Email Captured", 14),
        ("Email CVR", 11),
        ("Primary Conv", 13),
        ("Primary CVR", 11),
    ]
    add_header(ws1, cols1)
    ws1.row_dimensions[1].height = 28

    for i, r in enumerate(summary, 2):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        email_cvr = pct_val(r["email"], r["clicks"])
        primary_cvr = pct_val(r["primary"], r["clicks"])
        row_data = [r["url"], r["clicks"], round(r["email"], 1), email_cvr, round(r["primary"], 1), primary_cvr]
        for j, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=j, value=val)
            cell.fill = fill
            cell.border = BORDER
            if j in (4, 6) and val is not None:
                cell.number_format = "0.00%"
            if j in (2, 3, 5):
                cell.alignment = Alignment(horizontal="right")

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:F{len(summary) + 1}"

    # --- Sheet 2: Breakdown by Landing Page + Campaign ---
    ws2 = wb.create_sheet("By Landing Page + Campaign")
    cols2 = [
        ("Landing Page", 55),
        ("Campaign", 40),
        ("Clicks", 10),
        ("Email Captured", 14),
        ("Email CVR", 11),
        ("Primary Conv", 13),
        ("Primary CVR", 11),
    ]
    add_header(ws2, cols2)
    ws2.row_dimensions[1].height = 28

    current_url = None
    row_num = 2
    for r in by_url_campaign:
        is_new_url = r["url"] != current_url
        current_url = r["url"]
        email_cvr = pct_val(r["email"], r["clicks"])
        primary_cvr = pct_val(r["primary"], r["clicks"])
        row_data = [
            r["url"] if is_new_url else "",
            r["campaign"],
            r["clicks"],
            round(r["email"], 1),
            email_cvr,
            round(r["primary"], 1),
            primary_cvr,
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=j, value=val)
            cell.border = BORDER
            if is_new_url and j == 1:
                cell.fill = URL_FILL
                cell.font = Font(bold=True)
            if j in (5, 7) and val is not None:
                cell.number_format = "0.00%"
            if j in (3, 4, 6):
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:G{row_num - 1}"

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_XLS)
    print(f"Spreadsheet saved to: {OUTPUT_XLS}")


def main():
    run_date = datetime.utcnow().strftime("%B %d, %Y")

    print("Connecting to Google Ads API...")
    client = GoogleAdsClient.load_from_storage()
    ga = client.get_service("GoogleAdsService")

    acct_query = "SELECT customer.id, customer.descriptive_name FROM customer LIMIT 1"
    for row in ga.search(customer_id=CUSTOMER_ID, query=acct_query):
        print(f"Account: {row.customer.descriptive_name} ({row.customer.id})")

    print("Fetching clicks by landing page + campaign...")
    clicks_data = fetch_clicks(ga)

    print("Fetching conversions by landing page + campaign + action...")
    conv_data = fetch_conversions(ga)

    rows = build_rows(clicks_data, conv_data)
    summary = summarize_by_url(rows)
    by_url_campaign = summarize_by_url_campaign(rows)

    unique_urls = sorted({r["url"] for r in summary})
    print(f"  {len(unique_urls)} unique landing pages (after normalization), {len(by_url_campaign)} landing page / campaign combinations.")

    table_summary = md_table_summary(summary)
    table_by_campaign = md_table_by_url_campaign(by_url_campaign)

    report = f"""# Helcim — Landing Page Conversion Rate Report
*Generated: {run_date} · Date range: Last 30 days · Account: {CUSTOMER_ID}*
*URLs normalized: `{{ignore}}` stripped, query params removed, trailing slashes unified.*

---

## Unique Landing Pages ({len(unique_urls)})

{chr(10).join(f"- {u}" for u in unique_urls)}

---

## Summary by Landing Page

{table_summary}

---

## Breakdown by Landing Page + Campaign

{table_by_campaign}

---

*Email CVR = Conversion 1 (E-mail Captured) / clicks · Primary CVR = metrics.conversions (Account Approved) / clicks*
"""

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_MD, "w") as f:
        f.write(report)
    print(f"Markdown saved to: {OUTPUT_MD}")

    write_xlsx(summary, by_url_campaign, run_date)

    print("\n--- Summary by Landing Page ---")
    print(table_summary)


if __name__ == "__main__":
    main()
